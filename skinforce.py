import tkinter as tk
import tkinter.font
import tkinter.messagebox as msgbox
import random
import serial
import serial.tools.list_ports as sp
import threading
import json
import time
import openpyxl
import numpy as np

class ExperimentApp:

    jsonName, xlsxName, type = "data1.json", "exp1.xlsx", 3
    # jsonName, xlsxName, type = "data2.json", "exp2.xlsx", 1
    A, B = 1, 0

    def __init__(self, root):
        
        self.root = root
        self.root.title("Pseudoskin Experiment Program")

        self.width = self.root.winfo_screenwidth()
        self.height = self.root.winfo_screenheight()
        geometryString = f"{self.width}x{self.height}"
        self.root.geometry(geometryString)
        self.root.attributes("-fullscreen", True)
        self.root.resizable(True, True)

        #### JSON initialization #####

        try:
            with open(self.jsonName, 'r') as file:
                self.data = json.load(file)
        except FileNotFoundError:
            self.data = {}

        ######### variables ##########

        # initial settings
        self.handImages = {"knuckle": tk.PhotoImage(file="images/knuckle.png"),
                           "back of hand": tk.PhotoImage(file="images/backofhand.png"),
                           "forearm": tk.PhotoImage(file="images/forearm.png")}
        self.subjectName = ""
        self.bodyLocation = ""
        self.blockNumber = 0
        self.threshold = 30

        # calibration
        self.checkMaximum = False
        self.hapticOn = False
        self.prevForce = 0
        self.maximum = 0
        
        # experiment
        self.isRight = False       
        self.isTryStarted = True
        self.isPaused = False
        self.idx = 0
        self.scale = 0
        self.problemData = list()
        
        # results
        self.reactionTime = time.time()
        self.crossings = 0
        self.firstTouch = 0
        self.xlData = list()
        
        ##############################

        self.frames = []
        for _ in range(65):
            frame = tk.Frame(self.root)
            self.frames.append(frame)
        
        self.currentFrameIndex = 0    

        self.functionList = []
        self.functionList.append((self.initialPage, (0,)))

        self.showFrame()

        ##############################

        self.setSensorThread = threading.Thread(target=self.setSensor)
        self.setSensorThread.daemon = True  # It will be dead if the main thread is dead.
        self.setSensorThread.start()
        
        return

    def initialPage(self, page):
        
        frame = self.frames[page]

        titlefont= tkinter.font.Font(size=96, weight= "bold")
        buttonfont = tkinter.font.Font(size=72, weight= "normal")

        ###############################
        
        top = tk.Frame(frame)
        top.pack(side= "top", fill= "both", expand= True)

        label = tk.Label(top, text="피험자 번호를 입력해주세요.", font= titlefont)
        label.pack(fill= "both", expand= True)
        
        ###############################

        bottom = tk.Frame(frame)
        bottom.pack(side= "bottom", fill= "both", expand= True)

        entryFrame = tk.Frame(bottom)

        entry = tk.Entry(entryFrame, relief= "raised", font= buttonfont)
        entry.pack(side="left", padx= 20)
        entry.focus()

        checkButton = tk.Button(entryFrame, text="Check", command= lambda: self.checkDuplicate(str(entry.get()), nextButton), font= buttonfont, repeatinterval= 1000)
        checkButton.pack(side="right")

        entryFrame.pack(pady= 80)

        ###############################

        nextButton = tk.Button(bottom, text="Next", state= "disable", command=self.showNextFrame, font= buttonfont)
        nextButton.pack()

    def calibrationPage(self, target, page):

        self.bodyLocation = target
        self.hapticOn = False
        self.A, self.B = 1, 0
        frame = self.frames[page]

        titlefont= tkinter.font.Font(size= 96, weight= "bold")
        explainfont = tkinter.font.Font(size= 72, weight= "bold")
        buttonfont = tkinter.font.Font(size= 64, weight= "normal")

        ############ title ############
        
        top = tk.Frame(frame)
        top.pack(side= "top", fill= "both", expand= True)

        label = tk.Label(top, text="Calibration", font= titlefont)
        label.pack(fill= "both", expand= True)
        
        ############ image ############

        middle = tk.Frame(frame)
        middle.pack(fill= "both", expand= True)

        label = tkinter.Label(middle, image=self.handImages[target])
        label.pack(fill= "y", expand= True)

        ###############################

        bottom = tk.Frame(frame)
        bottom.pack(side= "bottom", fill= "both", expand= True)

        measureButton = tk.Button(bottom, text= "Measure", font= buttonfont, command= lambda: self.measurementPage(nextButton))
        measureButton.pack(side= "left", padx= 100)

        nextButton = tk.Button(bottom, text="Next", state= "disable", command=self.showNextFrame, font= buttonfont)
        nextButton.pack(side= "right", padx= 100)

        self.explainLabel = tk.Label(bottom, text= target + "\nPress the Measure Button.", font= explainfont)
        self.explainLabel.pack(fill= "y", expand= True)

    def measurementPage(self, button):
        explainFont = tkinter.font.Font(size= 84, weight= "normal")
        measurementFont = tkinter.font.Font(size= 128, weight= "bold")
        buttonfont = tkinter.font.Font(size= 64, weight= "normal")

        print("A, B값을 공백을 포함하여 입력해주세요.(ex. 2.4 3.6) :", end= " ")
        self.A, self.B = map(float, input().split())

        # whenever we measure, should initialize maximum value and turn on the checkMaximum.
        self.maximum = 0
        self.checkMaximum = True

        popUp = tk.Toplevel(self.root)
        popUp.title("Measurement Page (measuring maximum value)")
        geometryString = f"{self.width}x{self.height}"
        popUp.geometry(geometryString)

        explainLabel = tk.Label(popUp, text= "Press with as much force as possible without pain.", font= explainFont)
        explainLabel.place(relx= 0.5, rely= 0.25, anchor= "s")

        ##############################

        self.measurementLabel = tk.Label(popUp, text = "0.00 gf", font= measurementFont)
        self.measurementLabel.place(relx = 0.4, rely= 0.75, anchor= "se")

        self.maximumLabel = tk.Label(popUp, text= "0.00 gf", font= measurementFont, fg= "RoyalBlue3")
        self.maximumLabel.place(relx = 0.6, rely= 0.75, anchor= "sw")

        ##############################

        popUp.bind("<Button-3>", lambda event: self.measureAgain())

        saveButton = tk.Button(popUp, text= "Set", command= lambda: self.saveMaximumValue(popUp), font= buttonfont)
        saveButton.place(relx= 0.5, rely= 0.75, anchor= "ne")

        hapticButton = tk.Button(popUp, text = "Haptic", command= lambda: { self.Oser.write("b".encode('utf-8')) }, font= buttonfont)
        hapticButton.place(x= self.width - 100, y= self.height - 100, anchor= "se")

        if button != None:
            button.configure(state= "active")
        
        return
    
    def measureAgain(self):
        self.maximum = 0
        self.measurementLabel.config(text= "0.00 N")

    def saveMaximumValue(self, w):
        if self.maximum > 10:
            if hasattr(self, 'explainLabel'):
                self.explainLabel.config(text= "Measurement has been completed.\nPlease click the next button.")
            self.checkMaximum = False
            self.threshold = self.maximum / 32
            w.destroy()
        else:
            msgbox.showinfo("측정 오류", "값이 매우 작습니다.\n다시 눌러주세요.")
        
        return
 
    def preExperimentPage(self, page, haptics, state):

        if state != "Tutorial":
            self.blockNumber = (self.blockNumber + 1) % 4
            if self.blockNumber == 0: self.blockNumber = 1
        else:
            self.blockNumber = 0

        self.isRight = False
        self.firstTouch = 0

        titlefont= tkinter.font.Font(size=88, weight= "bold")
        hapticfont = tkinter.font.Font(size=72, weight= "bold")
        buttonfont = tkinter.font.Font(size=72, weight= "normal")
        
        frame = self.frames[page]
        top = tk.Frame(frame)
        top.pack(side= "top", fill= "both", expand= True)

        hapticLabel = tk.Label(top, text= "Haptic" if haptics else "No Haptic", font= hapticfont)
        hapticLabel.place(x = 100, y = 100)

        label = tk.Label(top, text= f"When you press the {state} start button, {state} starts.", font= titlefont)
        label.pack(anchor= "s", expand= True)

        bottom = tk.Frame(frame)
        bottom.pack(side= "bottom", fill= "both", expand= True)

        startButton = tk.Button(bottom, text=f"{state} Start", font=buttonfont, command=lambda: (self.showNextFrame(), setattr(self, "hapticOn", True if haptics else False), setattr(self, "firstTouch", 0), setattr(self, "isRight", False)))

        startButton.pack(pady= 120)

    def experimentPage(self, target, scale, page):
        
        self.scale = scale
        frame = self.frames[page]
        self.reactionTime = time.time()

        partFont = tkinter.font.Font(size= 72, weight = "bold")
        progressFont = tkinter.font.Font(size= 36, weight= "normal")

        ###############################

        whichPart = tk.Label(frame, text= target, font= partFont)
        whichPart.place(x= 100, y= 100)

        comment = "When you start moving the cursor\nPlease input as quickly and accurately as possible."
        explainComment = tk.Label(frame, text= comment, font= progressFont, justify= "left")
        explainComment.place(x= 100, y= 290, anchor= "sw")

        self.progressLabel = tk.Label(frame, text= str(self.idx), font= progressFont)
        self.progressLabel.place(x = 100, y = 310)
        
        ########### Canvas ############

        self.stickCanvas = tk.Canvas(frame, width= self.width / 10, height = self.height * 0.9)
        self.stickCanvas.place(relx=0.45, rely=0.05)

        self.stick = self.stickCanvas.create_rectangle(0, 0, self.width / 10, self.height * 0.9, fill= "snow", width= 0)

        x1, y1, x2, y2 = self.stickCanvas.coords(self.stick)

        self.gapY = (y2 - y1) / scale

        ranges = [i * 0.18 * y2 for i in range(1, 6)]
        self.problems = list()
        self.rects = list()
        for i in range(scale):
            addY = i * self.gapY
            for j in ranges:
                if addY <= j < addY + self.gapY:
                    rect = self.stickCanvas.create_rectangle(x1, y2 - addY - self.gapY, x2, y2 - addY, fill= "white smoke", width= 0)
                    self.rects.append(rect)
                    self.problems.append(y2 - addY - self.gapY)
                    break
            self.stickCanvas.create_line(x1, addY, x2, addY, fill= "light grey")

        self.now = self.stickCanvas.create_rectangle(0, self.height * 0.9, self.width / 10, self.height * 0.9 / scale * (scale + 1), fill= "light pink", width= 0)
        self.sensorBox = self.stickCanvas.create_rectangle(x1, y2, x2, y2 + self.gapY, fill= "pale violet red", width= 0)
        self.sensorLine = self.stickCanvas.create_line(x1, y2, x2, y2, width= self.gapY / 20, fill= "maroon")

        ###############################

        self.skipButton = tk.Button(frame, text= "Skip", command= lambda: (self.showNextFrame(), setattr(self, "idx", 0)), font= partFont)
        self.skipButton.place(x = self.width - 100, y = self.height - 210, anchor= "se")

        nextButton = tk.Button(frame, text= "Next", command= lambda: self.showProblems(scale), font= partFont)
        nextButton.place(x = self.width - 100, y = self.height - 190, anchor= "ne")

        self.problems = self.problems * 4
        random.shuffle(self.problems)
        
        self.showProblems(scale)
        self.stickCanvas.bind("<Button-3>", lambda event: self.showProblems(scale))
        frame.bind("<Button-3>", lambda event: self.showProblems(scale))

    def endPage(self, page):
        frame = self.frames[page]

        titlefont= tkinter.font.Font(size=96, weight= "bold")
        buttonfont = tkinter.font.Font(size=72, weight= "normal")
        
        top = tk.Frame(frame)
        top.pack(side= "top", fill= "both", expand= True)

        ###############################

        label = tk.Label(top, text= "The experiment is all over. thank you!", font= titlefont)
        label.pack(anchor= "s", expand= True)

        ###############################

        bottom = tk.Frame(frame)
        bottom.pack(side= "bottom", fill= "both", expand= True)

        endButton = tk.Button(bottom, text= "Finish", font= buttonfont, command= self.root.destroy)
        endButton.pack(pady= 120)

        ###############################


        self.data[self.subjectName]["progress"] = 1
        self.saveData(self.data)

        # need to write codes about analyzing and saving data.



    # In another threads, value of sensor always calculate and change the value.
    def setSensor(self):
        connected_ports = []
        for i in sp.comports():
            connected_ports.append(i.device)
        print(*connected_ports) 

        with serial.Serial() as self.Oser:
            self.Oser.port = connected_ports[3]
            self.Oser.baudrate = 9600
            self.Oser.timeout = 10
            self.Oser.write_timeout = 10
            self.Oser.open()
            while self.Oser.is_open:
                # while self.Oser.is_open:  
                with serial.Serial() as ser:
                    ser.port = connected_ports[2]
                    ser.baudrate = 9600
                    ser.timeout = 10
                    ser.write_timeout = 10
                    ser.open()

                    while ser.is_open:
                        ser.write("1".encode())
                        line = ser.readline()
                        if line:
                            try:
                                line = line.decode().strip()
                                initialValue = float(line)
                                self.value = round(self.A * (np.exp(self.B * initialValue) - 1), 3)

                                if abs(self.value - self.prevForce) >= self.threshold and self.hapticOn:
                                    # haptic reaction disposures when the task does not stop.
                                    if not self.isPaused:
                                        sig = "b"
                                        sig = sig.encode('utf-8')
                                        self.Oser.write(sig)
                                    self.prevForce = self.value

                                if self.checkMaximum and hasattr(self, 'measurementLabel') and hasattr(self, 'maximumLabel'):
                                    # if self.value > 0:
                                    self.maximum = max(self.value, self.maximum)
                                    self.root.after(0, self.updateMaximumLabel, (self.value, self.maximum))

                                if hasattr(self, 'stickCanvas') and hasattr(self, 'sensorLine'):
                                    self.moveSensorLine(self.scale)

                                if self.isTryStarted and self.value > 1:
                                    self.reactionTime = time.time()
                                    self.isTryStarted = False

                            except ValueError:
                                self.value = 0
 

    # This function shows realtime value and maximum value.
    def updateMaximumLabel(self, values):
        value, maximum = values
        self.measurementLabel.config(text= str(value) + " N")
        self.maximumLabel.config(text=str(maximum) + " N")
        self.crossings = 0
        self.isRight = False

    # This function moves line and box including line in realtime.
    def moveSensorLine(self, scale):      
        stickHeight = self.height * 0.9
        fraction = min(self.value / (self.maximum if self.maximum else 1500), 1)
        linePlace = stickHeight - stickHeight * fraction
        x1, _, x2, _ = self.stickCanvas.coords(self.sensorLine)

        for i in range(1, scale + 1):
            if (i - 1) * self.gapY <= linePlace < i * self.gapY:
                self.stickCanvas.coords(self.sensorBox, x1, (i - 1) * self.gapY, x2, i * self.gapY)
                break

        self.stickCanvas.coords(self.sensorLine, x1, linePlace, x2, linePlace)

        # check whether now the lineplace is in right place or not
        problemY = self.stickCanvas.coords(self.now)[1]

        self.problemData = [problemY, problemY + self.gapY, linePlace]

        if problemY <= linePlace < problemY + self.gapY:
            if not self.isRight:
                self.isRight = True
            if self.firstTouch == 0:
                self.firstTouch = time.time()

        else:
            if self.isRight:
                self.isRight = False
                self.crossings += 1


    def saveData(self, data):
        with open(self.jsonName, 'w') as file:
            json.dump(data, file, indent=4)       
  
    # This function checks duplicated subject number and sets the inital experiment environments.
    def checkDuplicate(self, num, button):
        self.subjectName = num
        self.progress = self.data[self.subjectName]["progress"]
        self.targets = self.data[self.subjectName]["targets"]
        self.haptics = self.data[self.subjectName]["haptics"]
        self.scales = self.data[self.subjectName]["scales"]
        
        if self.progress:
            print(f"이미 피험자 {self.subjectName}의 데이터가 존재합니다.")
            msgbox.showinfo("중복되는 번호", "다른 번호를 사용해주세요.")
            button.config(state= "disable")
        else:
            msgbox.showinfo("", "사용 가능한 번호입니다.")
            self.todoList = self.targets
            self.addContentToFrames()
            button.config(state= "active")
   
    # This function shows next problem.
    def showProblems(self, scale):
        
        if self.isPaused:
            return

        # escape condition
        if self.idx == 20 or (scale == self.idx == 5):
            if self.value < self.gapY / 2:
                return
            checkTime = time.time()
            self.idx = 0

            self.xlData.append([self.subjectName, self.bodyLocation, self.maximum, self.hapticOn, self.scale, self.blockNumber]
                               + self.problemData
                               + [self.isRight, self.value, self.crossings]
                               + list(map(lambda x: round(x % 86400, 3), [self.reactionTime, self.firstTouch, checkTime]))
                               )
            print("피험자 번호, 타겟 부위, 최대 힘, 햅틱 유무, 스케일, 블럭, 문제 앞, 문제 뒤, 현재 위치, 정오표, 현재 힘, 교차 수, 반응 시간, 처음 정답 시간, 입력 시간")
            print(*self.xlData[-1], sep= " ")

            self.saveXlData()

            self.showNextFrame()
            return
        
        # It does not implement at the first time.
        if self.idx:
            if self.value < self.gapY / 2:
                return
            checkTime = time.time()

            # Hide the current box
            self.stickCanvas.itemconfigure(self.now, state="hidden")
            self.stickCanvas.itemconfigure(self.sensorBox, state= "hidden")
            self.stickCanvas.itemconfigure(self.sensorLine, state= "hidden")
            self.skipButton.config(state= "disable")
            for i in self.rects:
                self.stickCanvas.itemconfig(i, state= "hidden")
            self.stickCanvas.itemconfig(self.stick, fill= "gray")
            self.isPaused = True
            
            # Schedule the function to show the box again after 5 seconds
            self.root.after(5000, self.showBoxAgain)

            self.xlData.append([self.subjectName, self.bodyLocation, self.maximum, self.hapticOn, self.scale, self.blockNumber]
                               + self.problemData
                               + [self.isRight, self.value, self.crossings]
                               + list(map(lambda x: round(x % 86400, 3), [self.reactionTime, self.firstTouch, checkTime]))
                               )
            print("피험자 번호, 타겟 부위, 최대 힘, 햅틱 유무, 스케일, 블럭, 문제 앞, 문제 뒤, 현재 위치, 정오표, 현재 힘, 교차 수, 반응 시간, 처음 정답 시간, 입력 시간")
            print(*self.xlData[-1], sep= " ")
            self.isRight = False

        else:
            self.nextBoxPlace()


    def nextBoxPlace(self):
        x1, _, x2, _ = self.stickCanvas.coords(self.now) 

        newY = self.problems[self.idx]
        self.stickCanvas.coords(self.now, x1, newY, x2, newY + self.gapY)

        self.idx += 1
        self.progressLabel.config(text= str(self.idx))

    def showBoxAgain(self):
        # Show the box again
        self.stickCanvas.itemconfigure(self.now, state="normal")
        self.stickCanvas.itemconfigure(self.sensorBox, state= "normal")
        self.stickCanvas.itemconfigure(self.sensorLine, state= "normal")
        self.skipButton.config(state= "active")
        for i in self.rects:
                self.stickCanvas.itemconfig(i, state= "normal")
        self.stickCanvas.itemconfig(self.stick, fill= "snow")
        # Call the function to move to the next box
        self.nextBoxPlace()
        self.isPaused = False
        self.isTryStarted = True
        self.firstTouch = 0
        self.crossings = 0

    def saveXlData(self):
        try:
            workbook = openpyxl.load_workbook(self.xlsxName)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            header = [
                "피험자 번호", "타겟 부위", "최대 힘", "햅틱 유무",
                "스케일", "블럭", "문제 앞", "문제 뒤",
                "현재 위치", "정오표", "현재 힘", "교차 수",
                "반응 시간", "처음 정답 시간", "입력 시간"
            ]
            sheet.append(header)

        last_row = sheet.max_row + 1

        for rowData in self.xlData:
            sheet.append(rowData)

        workbook.save(self.xlsxName)
        print("파일을 저장하였습니다.")
        self.xlData = []

    # In this function, we made pages to integrate.
    def addContentToFrames(self):
        page = 1
        print("이 순서대로 진행됩니다.")
        print(*self.todoList, sep= " ")
        print(*self.haptics, sep= " ")
        for i in self.todoList:
            self.functionList.append((self.calibrationPage, (i, page))); page += 1
            for j in self.haptics:
                self.functionList.append((self.preExperimentPage, (page, j, "Tutorial"))); page += 1
                self.functionList.append((self.experimentPage, (i, self.scales[0], page))); page += 1
                for k in self.scales:
                    for _ in range(self.type):
                        self.functionList.append((self.preExperimentPage, (page, j, "Experiment"))); page += 1
                        self.functionList.append((self.experimentPage, (i, k, page))); page += 1
        self.functionList.append((self.endPage, (page,)))

    # In this function, we add pages whenever we go to next page.
    def addPage(self, index):
        if 0 <= index < len(self.functionList):
            func, params = self.functionList[index]
            func(*params)

    def showFrame(self):
        self.addPage(self.currentFrameIndex)
        frame = self.frames[self.currentFrameIndex]
        frame.pack(fill=tk.BOTH, expand=True)

    def showNextFrame(self):
        # right after subjects confirm their input, cannot allow them to go to next frame.
        if self.isPaused:
            return
        self.frames[self.currentFrameIndex].pack_forget()
        self.currentFrameIndex = (self.currentFrameIndex + 1) % len(self.frames)
        self.showFrame()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExperimentApp(root)
    root.mainloop()
